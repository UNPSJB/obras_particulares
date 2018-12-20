#-*- coding:utf-8 -*-

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import  login_required
from .forms import *
from django.contrib import messages
from tipos.forms import *
from obras_particulares.views import *
from tramite.forms import FormularioIniciarTramite
from documento.forms import FormularioDocumentoSetFactory,FormularioCorreccionesDocumento
from documento.forms import metodo
from tramite.models import *
from django.core.mail import send_mail
from persona.models import *
from tramite.models import Tramite, Estado
import re
from datetime import datetime, date, time, timedelta
from dateutil import parser
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import View
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Table, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics import renderPDF
from reportlab.graphics.charts.barcharts import VerticalBarChart
import collections
from django.utils import timezone
from django.http import JsonResponse
from tipos.models import *
from django.db.models import F, Q, When
import pandas as pd
import operator as operator
from operator import attrgetter

'''propietario ------------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('propietario')
def mostrar_propietario(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        "ctxtramitespropietario": listado_tramites_propietario(request),
        "ctxtramitesprofesionalpropietario": listado_tramites_cambiar_profesional(request)
    }
    for form_name, submit_name in FORMS_PROPIETARIO:
        KlassForm = FORMS_PROPIETARIO[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/propietario/propietario.html', values)

FORMS_PROPIETARIO = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def listado_tramites_propietario(request):
    pk_persona = request.user.persona.pk
    tramites = Tramite.objects.all()
    tramites_de_propietario = filter(lambda t: (t.propietario.persona.pk == pk_persona), tramites)
    tipos_ag = [11, 21, 28]  # agendados para inspeccion
    dia_hoy = date.today()
    tramites_no_aprobado_o_no_final_o_inspeccion = filter(lambda t: ((datetime.datetime.strftime(t.estado().timestamp, '%d/%m/%Y') == datetime.datetime.strftime(dia_hoy, '%d/%m/%Y') and
                                                                      (t.estado().tipo == tipos_ag[0] or
                                                                       t.estado().tipo == tipos_ag[1] or
                                                                       t.estado().tipo == tipos_ag[2])) or
                                                                     (str(t.estado()) != 'NoFinalizado' or
                                                                      str(t.estado()) != 'NoAprobado')), tramites_de_propietario)

    contexto = {'tramites': tramites_de_propietario, 'tramites_no_aprobado_o_no_final': len(list(tramites_no_aprobado_o_no_final_o_inspeccion))}
    return contexto


def cargar_aprobacion_propietario(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.SOLICITAR_APROBAR_TRAMITE_PROPIETARIO)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "aprobar_tramite" in request.POST:
                aprobar_tramite_propietario(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/propietario/cargar_aprobacion.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('propietario')


def aprobar_tramite_propietario(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(tramite.SOLICITAR_APROBAR_TRAMITE_PROPIETARIO, request.user)
        for docForm in documento_set:
                docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'Solicitud de aprobar tramite realizada.')
    except:
        messages.add_message(request, messages.ERROR, 'No se puede solicitar aprobar tramite para este tramite.')
    finally:
        return redirect('propietario')


def cargar_final_de_obra_total_propietario(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.SOLICITAR_FINAL_OBRA_TOTAL_PROPIETARIO)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "aprobar_final_de_obra_total" in request.POST:
                propietario_solicita_final_obra(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/propietario/cargar_final_de_obra_total.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('propietario')


def propietario_solicita_final_obra(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(Tramite.SOLICITAR_FINAL_OBRA_TOTAL_PROPIETARIO, request.user)
        for docForm in documento_set:
                docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'Final de obra total solicitado.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra total para ese tramite.')
    finally:
        return redirect('propietario')


def ver_historial_tramite(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    pk = int(pk_tramite)
    estados = Estado.objects.all()
    estados_de_tramite = filter(lambda e: (e.tramite.pk == pk), estados)
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = [];
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"))
    return render(request, 'persona/propietario/ver_historial_tramite.html', {"tramite": contexto0,
                                                                              "estadosp": contexto1,
                                                                              "fecha":fechas_del_estado,
                                                                              "perfil": perfil})


def documentos_de_estado(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e:(datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto= {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/propietario/documentos_de_estado.html', contexto)

"""
Metodo que se encarga de devolver un listado con los tramites correspondientes
para que el propietario los pueda cambar de profesional si lo desea
"""
def listado_tramites_cambiar_profesional(request):
    pk_persona = request.user.persona.pk
    tramites = Tramite.objects.all()
    tramites_de_propietario = filter(lambda t: (t.propietario.persona.pk == pk_persona
                                                and str(t.estado()) != 'NoAprobadoSolicitado'
                                                and str(t.estado()) != 'NoAprobado'
                                                and str(t.estado()) != 'AprobadoSolicitadoPorPropietario'
                                                and str(t.estado()) != 'NoFinalObraTotalSolicitado'
                                                and str(t.estado()) != 'AgendadoInspeccionFinal'
                                                and str(t.estado()) != 'InspeccionFinal'
                                                and str(t.estado()) != 'FinalObraTotalSolicitadoPorPropietario'
                                                and str(t.estado()) != 'Finalizado'
                                                and str(t.estado()) != 'Baja'), tramites)
    return tramites_de_propietario


def cambiar_profesional_de_tramite(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    usuarios = Usuario.objects.all()
    profesionales = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(lista)):
            if lista[i] == 'profesional':
                if u not in profesionales and u != tramite.profesional.persona.usuario:
                    profesionales.append(u)
    if request.method == "POST" and "cambiar_profesional" in request.POST:
        if request.POST["idempleado"]:
            pk_profesional = int(request.POST["idempleado"])
            print(request.POST['idempleado'])
            print(pk_profesional)
            profesional = Persona.objects.get(pk=pk_profesional)
            if tramite.profesional.persona.id != profesional.id:
                tramite.cambiar_profesional(profesional.profesional)
                messages.add_message(request, messages.SUCCESS, "El profesional del tramite ha sido cambiado")
            else:
                messages.add_message(request, messages.ERROR, "El profesional del tramite no ha sido cambiado. Selecciono el mismo profesional.")
        else:
            messages.add_message(request, messages.ERROR,"El profesional del tramite no ha sido cambiado. No se selecciono un profesional.")
    else:
        return render(request, 'persona/propietario/cambiar_profesional_de_tramite.html', {'tramite': tramite, "perfil": perfil, 'profesionales': profesionales})
    return redirect('propietario')


'''profesional -------------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('profesional')
def mostrar_profesional(request):
    usuario = request.user
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.INICIAR)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    tramite_form = FormularioIniciarTramite(initial={'profesional':usuario.persona.profesional.pk})
    propietario_form = FormularioPropietario()
    propietario = None
    if request.method == "POST" and "propietario" in request.POST:
        tramites = Tramite.objects.all()
        t = filter(lambda tramite: (tramite.domicilio == request.POST['domicilio']), tramites)
        if int(usuario.persona.profesional.categoria) <= int(request.POST['tipo_obra']) and len(list(t)) == 0:
            personas = Persona.objects.filter(dni=request.POST["propietario"])
            persona = personas.exists() and personas.first() or None
            documento_set = FormularioDocumentoSet(request.POST, request.FILES)
            propietario_form = FormularioPropietario(request.POST)
            tramite_form = FormularioIniciarTramite(request.POST)
            documento_set = FormularioDocumentoSet(request.POST, request.FILES)
            propietario = propietario_form.obtener_o_crear(persona)
            if propietario is not None and tramite_form.is_valid() and documento_set.is_valid():
                tramite = tramite_form.save(propietario=propietario, commit=False)
                lista = []
                for docForm in documento_set:
                    lista.append(docForm.save(commit=False))
                Tramite.new(usuario, propietario, usuario.persona.profesional, request.POST['tipo_obra'], request.POST['medidas'], request.POST['domicilio'], lista, request.POST['destino_obra'])
                tramite_form = FormularioIniciarTramite(initial={'profesional':usuario.persona.profesional.pk})
                propietario_form = None
                messages.add_message(request, messages.SUCCESS,'Solicitud de iniciar tramitre realizada con exito.')
            else:
                messages.add_message(request, messages.ERROR, 'Propietario no existe, debe darlo de alta para iniciar al tramite.')
        else:
            propietario_form = None
            if int(usuario.persona.profesional.categoria) > int(request.POST['tipo_obra']):
                messages.add_message(request, messages.ERROR,
                                     'Categoria del profesional no es suficiente para el tipo de tramite que desea iniciar.')
            else:
                messages.add_message(request, messages.ERROR,
                                     'Ya existe un tramite para este domicilio.')
    else:
        propietario_form = None
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        'documentos_requeridos': tipos_de_documentos_requeridos,
        'ctxtramitesprofesional': listado_tramites_de_profesional(request),
        'tramite_form': tramite_form,
        'propietario_form': propietario_form,
        'documento_set': documento_set,
        'ctxtramcorregidos': tramites_corregidos(request)
    }
    for form_name, submit_name in FORMS_ADMINISTRATIVO:
        KlassForm = FORMS_ADMINISTRATIVO[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS,"La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/profesional/profesional.html', values)

FORMS_PROFESIONAL = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def listado_tramites_de_profesional(request):
    tramites = Tramite.objects.all()
    personas = Persona.objects.all()
    usuario = request.user
    lista_de_persona_que_esta_logueada = filter(lambda persona: (persona.usuario is not None and persona.usuario == usuario), personas)
    persona = list(lista_de_persona_que_esta_logueada).pop()
    profesional = persona.get_profesional()
    tramites_de_profesional = filter(lambda tramite: (tramite.profesional == profesional), tramites)
    tipos_ag = [11, 21, 28]  # agendados para inspeccion
    tipo_ip = 12  # primer inspeccion
    tramites_inspecion_dia = filter(lambda t: ((datetime.datetime.strftime(t.estado().timestamp, '%d/%m/%Y') == datetime.datetime.strftime(datetime.datetime.now(), '%d/%m/%Y') and
                                               (t.estado().tipo == tipos_ag[0] or
                                                t.estado().tipo == tipos_ag[1] or
                                                t.estado().tipo == tipos_ag[2])) or (t.estado().tipo == tipo_ip and t.monto_pagado and t.monto_pagado >= (t.monto_a_pagar/12))), tramites_de_profesional)
    contexto = {'tramites_de_profesional': tramites_de_profesional, 'tramites_inspeccion_dia': len(list(tramites_inspecion_dia))}
    return contexto


def tramites_corregidos(request):
    tramites = Tramite.objects.all()
    personas = Persona.objects.all()
    usuario = request.user
    lista_de_persona_que_esta_logueada = filter(lambda persona: (persona.usuario is not None and persona.usuario == usuario), personas)
    persona = list(lista_de_persona_que_esta_logueada).pop()
    profesional = persona.get_profesional()
    tramites_de_profesional = filter(lambda tramite: (tramite.profesional == profesional), tramites)
    tram_corregidos = filter(lambda tramite: (str(tramite.estado()) == 'ConCorrecciones' or
                                              str(tramite.estado()) == 'ConCorreccionesDeVisado' or
                                              str(tramite.estado()) == 'ConCorreccionesDePrimerInspeccion' or
                                              str(tramite.estado()) == 'ConCorreccionesDeInspeccion' or
                                              str(tramite.estado()) == 'ConCorreccionesDeInspeccionFinal'),
                             tramites_de_profesional)
    contexto = {'tramites': tram_corregidos, 'len_tramites': len(list(tram_corregidos))}
    return contexto


def ver_documentos_tramite_profesional(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    pk = int(pk_tramite)
    estados = Estado.objects.all()
    estados_de_tramite = filter(lambda e: (e.tramite.pk == pk), estados)
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = []
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"))
    return render(request, 'persona/profesional/vista_de_documentos.html', {"tramite": contexto0,
                                                                            "estadosp": contexto1,
                                                                            "fecha":fechas_del_estado,
                                                                            "perfil": perfil})


def profesional_solicita_aprobar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        if str(tramite.estado()) == 'PrimerInspeccion' and str(tramite.estado()) != 'NoAprobado':
            tramite.hacer(Tramite.SOLICITAR_APROBAR_TRAMITE, request.user)
            messages.add_message(request, messages.SUCCESS, 'Solicitud de aprobar tramite realizada.')
        else:
            messages.add_message(request, messages.ERROR, 'No puede solicitar aprobar tramite para ese tramite.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar aprobar tramite para ese tramite.')
    finally:
        return redirect('profesional')


def cargar_no_aprobar_profesional(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.SOLICITAR_NO_APROBAR_TRAMITE)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "no_aprobar_tramite" in request.POST:
                profesional_solicita_no_aprobar_tramite(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/profesional/cargar_no_aprobacion.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('profesional')


def profesional_solicita_no_aprobar_tramite(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(Tramite.SOLICITAR_NO_APROBAR_TRAMITE, request.user)
        for docForm in documento_set:
            docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'Solicitud de no aprobar tramite realizada.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar no aprobar tramite para ese tramite.')
    finally:
        return redirect('profesional')


def cargar_final_de_obra_total_profesional(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.SOLICITAR_FINAL_OBRA_TOTAL)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "aprobar_final_de_obra_total" in request.POST:
                profesional_solicita_final_obra(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/profesional/cargar_final_de_obra_total.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('profesional')


def profesional_solicita_final_obra(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        if str(tramite.estado()) != 'NoFinalizado':
            tramite.hacer(Tramite.SOLICITAR_FINAL_OBRA_TOTAL, request.user)
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            messages.add_message(request, messages.SUCCESS, 'Final de obra solicitado.')
        else:
            messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra para ese tramite.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra para ese tramite.')
    finally:
        return redirect('profesional')


def cargar_no_final_de_obra_total_profesional(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.SOLICITAR_NO_FINAL_OBRA_TOTAL)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "no_aprobar_final_de_obra_total" in request.POST:
                profesional_solicita_no_final_obra(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/profesional/cargar_no_final_de_obra_total.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('profesional')


def profesional_solicita_no_final_obra(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(Tramite.SOLICITAR_NO_FINAL_OBRA_TOTAL, request.user)
        for docForm in documento_set:
                docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'No Final de obra solicitado.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el no final de obra para ese tramite.')
    finally:
        return redirect('profesional')


def profesional_solicita_final_obra_parcial(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        estados = Estado.objects.all()
        tipo = 17
        estado = filter(lambda e: (e.tipo == tipo and str(e.tramite.pk) == str(pk_tramite)), estados)
        if len(list(estado)) == 0:
            tramite.hacer(Tramite.SOLICITAR_FINAL_OBRA_PARCIAL, request.user)
            messages.add_message(request, messages.SUCCESS, 'Final de obra parcial solicitado.')
        else:
            messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra parcial para ese tramite.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra parcial para ese tramite.')
    finally:
        return redirect('profesional')


from documento.models import Documento
def crear_correcciones(request, tramite, tipos_correcciones):
    files = request.FILES.getlist('file_field')
    for f in files:
        doc = Documento.objects.create(tipo_documento=tipos_correcciones[0], tramite=tramite, file=f)


def ver_documentos_corregidos(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.INGRESAR_CORRECCIONES)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    formCorrecciones = FormularioCorreccionesDocumento

    documento_set = FormularioDocumentoSet(initial=inicial)
    if request.method == "POST" and "enviar_correcciones" in request.POST:
        '''
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            enviar_correcciones(request, pk_tramite)
        '''
        crear_correcciones(request, tramite, tipos_de_documentos_requeridos)
        enviar_correcciones(request, pk_tramite)
    else:
        return render(request, 'persona/profesional/ver_documentos_corregidos.html', {'tramite': tramite,
                                                                                      "perfil": perfil,
                                                                                      'ctxdocumentoset': documento_set,
                                                                                      'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                                      'form_correcciones': formCorrecciones
                                                                                      })
    return redirect('profesional')


def enviar_correcciones(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.INGRESAR_CORRECCIONES, request.user)
    messages.add_message(request, messages.SUCCESS, 'Tramite con documentos corregidos y enviados')
    return redirect('profesional')


def documento_de_estado(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e:(datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto= {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/profesional/documento_de_estado.html', contexto)


'''administrativo ---------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('administrativo')
def mostrar_administrativo(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        "ctxprofesional": profesional_list(),
        "ctxpropietario": propietario_list(),
        "ctxtramitesiniciados": listado_de_tramites_iniciados(),
        "ctxtramitescorregidos": tramite_corregidos_list(),
        "ctxsolicitudesfinalobra": solicitud_final_obra_list(),
        "ctxsolicitudesaprobacion": solicitud_aprobacion_list(),
        "ctxsolicitudesnoaprobacion": solicitud_no_aprobacion_list(),
        "ctxpago": registrar_pago_tramite(request),
        "ctxtramitesvencidos": listado_tramites_vencidos()
    }
    for form_name, submit_name in FORMS_ADMINISTRATIVO:
        KlassForm = FORMS_ADMINISTRATIVO[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/administrativo/administrativo.html', values)


FORMS_ADMINISTRATIVO = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioGrupo,
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def profesional_list():
    personas = Persona.objects.all()
    profesionales = filter(lambda persona: (persona.usuario is None and persona.profesional is not None), personas)
    contexto = {'personas': profesionales,
                'len_personas': len(list(profesionales))
                }
    return contexto


def propietario_list():
    propietarios = Propietario.objects.all()
    propietarios_sin_usuario = filter(lambda propietario: (propietario.persona.usuario is None and propietario.persona is not None ), propietarios)
    contexto = {'propietarios': propietarios_sin_usuario,
                'len_propietarios': len(list(propietarios_sin_usuario))
                }
    return contexto


def listado_de_tramites_iniciados():
    argumentos = [Iniciado, ConCorreccionesRealizadas]
    tramites = Tramite.objects.en_estado(argumentos)
    contexto = {'tramites': tramites}
    return contexto


def tramite_corregidos_list():
    tramites = Tramite.objects.all()
    contexto = {'tramites': tramites}
    return contexto


def solicitud_final_obra_list():
    argumentos = [InspeccionFinal, FinalObraTotalSolicitadoPorPropietario]
    tramites = Tramite.objects.en_estado(argumentos)
    contexto = {'tramites': tramites}
    return contexto


def solicitud_aprobacion_list():
    argumentos = [AprobadoSolicitado, AprobadoSolicitadoPorPropietario]
    tramites = Tramite.objects.en_estado(argumentos)
    contexto = {'tramites': tramites}
    return contexto


def solicitud_no_aprobacion_list():
    tramites = Tramite.objects.en_estado(NoAprobadoSolicitado)
    contexto = {'tramites': tramites}
    return contexto


def listado_tramites_vencidos():
    argumentos_ap = [Iniciado, Aceptado, AgendadoParaVisado, Visado, AgendadoPrimerInspeccion, PrimerInspeccion,
                  NoAprobadoSolicitado, NoAprobado]
    tramites_ap = Tramite.objects.en_estado(argumentos_ap)
    argumentos_fo = [AprobadoSolicitado, Aprobado, NoAprobadoSolicitado, NoAprobado, AprobadoSolicitadoPorPropietario,
                  AprobadoPorPropietario, AgendadoInspeccion, Inspeccionado, FinalObraParcialSolicitado]
    tramites_fo = Tramite.objects.en_estado(argumentos_fo)
    estados = Estado.objects.all()
    tipo = 1
    estados_iniciado = filter(lambda e: (e.tipo == tipo), estados)
    tramites_vencidos = []
    for t in tramites_ap:
        for e in estados_iniciado:
            if e.tramite == t and (e.timestamp + timedelta(days=60)).strftime("%Y/%m/%d") < datetime.datetime.now().strftime("%Y/%m/%d"):
                tramites_vencidos.append(t)
    tramites_vencidos_no_pagados_no_renovados = []
    for tr in tramites_vencidos:
        if not tr.monto_pagado or tr.monto_pagado < (tr.monto_a_pagar / 12):
            tramites_vencidos_no_pagados_no_renovados.append(tr)
    tipoFOPS = 17
    for t in tramites_fo:
        estado_t = filter(lambda e: (e.tipo == tipoFOPS and str(e.tramite.pk) == str(t.pk)), estados)
        for e in estados_iniciado:
            if e.tramite == t and len(list(estado_t)) == 0 and (e.timestamp + timedelta(days=1095)).strftime("%Y/%m/%d") < datetime.datetime.now().strftime("%Y/%m/%d"):
                tramites_vencidos_no_pagados_no_renovados.append(t)
            elif e.tramite == t and len(list(estado_t)) > 0 and (e.timestamp + timedelta(days=1825)).strftime("%Y/%m/%d") < datetime.datetime.now().strftime("%Y/%m/%d"):
                tramites_vencidos_no_pagados_no_renovados.append(t)
    contexto = {'tramites': tramites_vencidos_no_pagados_no_renovados, 'tramites_vencidos_no_pagados_no_renovados': len(tramites_vencidos_no_pagados_no_renovados)}
    return contexto


def registrar_pago_tramite(request):
    if request.method == "POST":
        archivo_pago_form = FormularioArchivoPago(request.POST, request.FILES)
        if archivo_pago_form.is_valid():
            Pago.procesar_pagos(request.FILES['pagos'])
            messages.add_message(request, messages.SUCCESS, 'Se ha registrado el pago')
    else:
        archivo_pago_form = FormularioArchivoPago()
    return archivo_pago_form


def crear_usuario(request, pk_persona):
    usuario = request.user
    persona = get_object_or_404(Persona, pk=pk_persona)
    creado, password, usuario_creado = persona.crear_usuario()
    print(password)
    print(usuario_creado.username)
    if creado:
        messages.add_message(request, messages.SUCCESS, 'Profesional fue aceptado y su usuario creado.')
        send_mail(
            'Usuario habilitado',
            'Usted ya puede acceder al sistema: Nombre de usuario: '+persona.mail+' password: '+password,
            'infosopunpsjb@gmail.com',
            [persona.mail],
            fail_silently=False,
        )
    else:
        print("Mando correo informando que se cambio algo en su cuenta de usuario")
    return redirect(usuario.get_view_name())


def cargar_final_de_obra_total(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    estados = Estado.objects.all()
    tipoFOS = 23
    tipos_de_documentos_requeridos = []
    estadoFOS = filter(lambda e: (e.tipo == tipoFOS and str(e.tramite.pk) == str(pk_tramite)), estados)
    if len(list(estadoFOS)) > 0:
        tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.FINALIZAR)
    tipoNFOS = 25
    estadoNFOS = filter(lambda e: (e.tipo == tipoNFOS and str(e.tramite.pk) == str(pk_tramite)), estados)
    if len(list(estadoNFOS)) > 0:
        tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.NO_FINALIZAR)
    if str(tramite.estado()) == 'FinalObraTotalSolicitadoPorPropietario':
        tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.FINALIZAR)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            if "aprobar_final_de_obra_total" in request.POST:
                habilitar_final_obra(request, pk_tramite)
    else:
        return render(request, 'persona/administrativo/cargar_final_de_obra_total.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('administrativo')


def habilitar_final_obra(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        estados = Estado.objects.all()
        tipoFOS = 23
        estadoFOS = filter(lambda e: (e.tipo == tipoFOS and str(e.tramite.pk) == str(pk_tramite)), estados)
        if len(list(estadoFOS)) > 0:
            tramite.hacer(tramite.FINALIZAR, request.user)
            messages.add_message(request, messages.SUCCESS, 'Final de obra aprobado.')
        tipoNFOS = 25
        estadoNFOS = filter(lambda e: (e.tipo == tipoNFOS and str(e.tramite.pk) == str(pk_tramite)), estados)
        if len(list(estadoNFOS)) > 0 and str(tramite.estado()) != 'FinalObraTotalSolicitadoPorPropietario':
            tramite.hacer(tramite.NO_FINALIZAR, request.user)
            messages.add_message(request, messages.SUCCESS, 'No Final de obra aprobado.')
        if len(list(estadoNFOS)) > 0 and str(tramite.estado()) == 'FinalObraTotalSolicitadoPorPropietario':
            tramite.hacer(tramite.FINALIZAR, request.user)
            messages.add_message(request, messages.SUCCESS, 'Final de obra total solicitado por propietario aprobado.')
    except:
        messages.add_message(request, messages.ERROR, 'No se puede otorgar final de obra total para ese tramite.')
    finally:
        return redirect('administrativo')

def cargar_aprobacion(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.APROBAR_TRAMITE)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            if "aprobar_tramite" in request.POST:
                aprobar_tramite(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/administrativo/cargar_aprobacion.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('administrativo')


def aprobar_tramite(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(tramite.APROBAR_TRAMITE, request.user)
        for docForm in documento_set:
                docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'Tramite aprobado.')
    except:
        messages.add_message(request, messages.ERROR, 'No se puede aprobar este tramite.')
    finally:
        return redirect('administrativo')


def cargar_no_aprobacion(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.NO_APROBAR_TRAMITE)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    print("-----------------GET----------------")
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)

        if documento_set.is_valid():
            if "no_aprobar_tramite" in request.POST:
                no_aprobar_tramite(request, pk_tramite, documento_set)
    else:
        return render(request, 'persona/administrativo/cargar_no_aprobacion.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('administrativo')


def no_aprobar_tramite(request, pk_tramite, documento_set):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(tramite.NO_APROBAR_TRAMITE, request.user)
        for docForm in documento_set:
                docForm.save(tramite=tramite)
        messages.add_message(request, messages.SUCCESS, 'Tramite no aprobado.')
    except:
        messages.add_message(request, messages.ERROR, 'No se puede no aprobar este tramite.')
    finally:
        return redirect('administrativo')


def aceptar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.ACEPTAR, request.user)
    messages.add_message(request, messages.SUCCESS, "Tramite aceptado")
    return redirect('administrativo')


def rechazar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.CORREGIR, request.user, request.GET["msg"])
    messages.add_message(request, messages.SUCCESS, 'Tramite rechazado.')
    return redirect('administrativo')


def dar_baja_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    print(tramite.estado().tipo )
    if tramite.estado().tipo < 19:
        obs = "Tramite dado de baja. Se ha vencido el plazo de pago del permiso de construccion"
    else:
        obs = "Tramite dado de baja. Se ha vencido el plazo de construccion de la obra"
    tramite.hacer(tramite.DAR_DE_BAJA, request.user, obs)
    messages.add_message(request, messages.SUCCESS, obs)
    return redirect('administrativo')


def ver_un_certificado(request, pk_persona):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    persona = get_object_or_404(Persona, pk=pk_persona)
    return render(request, 'persona/administrativo/ver_certificado_profesional.html', {'persona': persona,
                                                                                       "perfil": perfil})


def ver_documentos_tramite_administrativo(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    return render(request, 'persona/administrativo/vista_de_documentos_administrativo.html', {'tramite': tramite,
                                                                                              "perfil": perfil})


"""
Metodo que se encarga de devolver todos los profesionales que tienen un usuario
Se utiliza en la vista de administrativo
"""
@login_required(login_url="login")
@grupo_requerido('administrativo')
def listado_profesionales_administrativo(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    personas = Persona.objects.all()
    profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
    contexto = {'profesionales': profesionales_con_usuario, "perfil": perfil}
    return render(request, 'persona/administrativo/profesional_list_con_usuario_administrativo.html', contexto)


"""
Metodo que se encarga de devolver todos los propietarios con usuario
Se utiliza en la vista de administrativo
"""
@login_required(login_url="login")
@grupo_requerido('administrativo')
def listado_propietarios_administrativo(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    propietarios = Propietario.objects.all()
    propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None ), propietarios)
    contexto = {'propietarios': propietarios_con_usuario, "perfil": perfil}
    return render(request, 'persona/administrativo/propietario_list_con_usuario_administrativo.html', contexto)


"""
Se encarga de devolver todos los profesionales con usuario en un archivo de excel
Se utiliza en la vista de administrativo
"""
class ReporteProfesionalesAdministrativoExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        cont = 0
        personas = Persona.objects.all()
        profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LISTADO DE PROFESIONALES'
        ws.merge_cells('B1:K1')
        ws['B2'] = 'NOMBRE'
        ws['C2'] = 'APELLIDO'
        ws['D2'] = 'MAIL'
        ws['E2'] = 'DIRECCION'
        ws['F2'] = 'CUIL'
        ws['G2'] = 'TELEFONO'
        ws['H2'] = 'PROFESION'
        ws['I2'] = 'CATEGORIA'
        ws['J2'] = 'MATRICULA'
        cont = 3
        for p in profesionales_con_usuario:
            ws.cell(row=cont, column=2).value = str(p.nombre)
            ws.cell(row=cont, column=3).value = str(p.apellido)
            ws.cell(row=cont, column=4).value = str(p.mail)
            ws.cell(row=cont, column=5).value = str(p.domicilio_persona)
            ws.cell(row=cont, column=6).value = str(p.cuil)
            ws.cell(row=cont, column=7).value = str(p.telefono)
            ws.cell(row=cont, column=8).value = str(p.profesional.profesion)
            ws.cell(row=cont, column=9).value = str(p.profesional.categoria)
            ws.cell(row=cont, column=10).value = str(p.profesional.matricula)
            cont = cont + 1
        nombre_archivo = "ListadoProfesionalesAdministrativoExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


"""
Se encarga de devolver todos los propietarios con usuario en un archivo de excel
Se utiliza en la vista de administrativo
"""
class ReportePropietariosAdministrativoExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        cont = 0
        propietarios = Propietario.objects.all()
        propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None), propietarios)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LISTADO DE PROPIETARIOS'
        ws.merge_cells('B1:G1')
        ws['B2'] = 'NOMBRE'
        ws['C2'] = 'APELLIDO'
        ws['D2'] = 'CUIL'
        ws['E2'] = 'TELEFONO'
        ws['F2'] = 'EMAIL'
        cont = 3
        for p in propietarios_con_usuario:
            ws.cell(row=cont, column=2).value = str(p.persona.nombre)
            ws.cell(row=cont, column=3).value = str(p.persona.apellido)
            ws.cell(row=cont, column=4).value = str(p.persona.cuil)
            ws.cell(row=cont, column=5).value = str(p.persona.telefono)
            ws.cell(row=cont, column=6).value = str(p.persona.mail)
            cont = cont + 1
        nombre_archivo = "ListadoPropietariosAdministrativoExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


"""
Se encarga de devolver todos los profesionales con usuario en un archivo de pdf
Se utiliza en la vista de administrativo
"""
class ReporteProfesionalesAdministrativoPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Listado de profesionales Administrativo.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )
        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))
        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)
        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Listado de profesionales'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))

        encabezados = ('NOMBRE', 'APELLIDO', 'MAIL', 'CUIL', 'TELEFONO', 'PROFESION', 'CAT.', 'MAT.')
        personas = Persona.objects.all()
        profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
        detalles = [
            (p.nombre, p.apellido, p.mail, p.cuil, p.telefono, p.profesional.profesion, p.profesional.categoria, p.profesional.matricula)
            for p in profesionales_con_usuario]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 4 * cm, 3 * cm, 2 * cm, 3.5 * cm, 1 * cm, 2.5 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response


"""
Se encarga de devolver todos los propietarios con usuario en un archivo de pdf
Se utiliza en la vista de administrativo
"""
class ReportePropietariosAdministrativoPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Listado de propietarios Administrativo.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )
        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))
        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)
        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Listado de propietarios'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))

        encabezados = ('NOMBRE', 'APELLIDO', 'CUIL', 'TELEFONO', 'EMAIL')
        propietarios = Propietario.objects.all()
        propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None),propietarios)
        detalles = [(p.persona.nombre, p.persona.apellido, p.persona.cuil, p.persona.telefono, p.persona.mail)
            for p in propietarios_con_usuario]
        detalle_orden = Table([encabezados] + detalles, colWidths=[3 * cm, 3 * cm, 3 * cm, 3 * cm, 5 * cm])

        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response


'''visador -----------------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('visador')
def mostrar_visador(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        "ctxtramaceptado": tramites_aceptados(),
        "ctxtramagendado": tramites_agendados(request),
        "ctxtramvisados": tramites_visados(request),
    }
    for form_name, submit_name in FORMS_VISADOR:
        KlassForm = FORMS_VISADOR[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/visador/visador.html', values)

FORMS_VISADOR = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioGrupo,
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def tramites_aceptados():
    argumentos = [Aceptado, CorreccionesDeVisadoRealizadas]
    aceptados = Tramite.objects.en_estado(argumentos)
    contexto = {'tramites': aceptados}
    return contexto


def agendar_tramite_para_visado(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.AGENDAR_VISADO, request.user)
    messages.add_message(request, messages.SUCCESS, "El visado ha sido agendado")
    return redirect('visador')


def tramites_agendados(request):
    usuario = request.user
    argumentos = [AgendadoParaVisado]
    agendados = Tramite.objects.en_estado(argumentos)
    tramites_del_visador = list(filter(lambda t: t.estado().usuario == usuario, agendados))
    contexto = {'tramites': tramites_del_visador, 'len_tramites_del_visador': len(list(tramites_del_visador))}
    return contexto


def tramites_visados(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipos = [5, 8]
    estados_visado = filter(lambda estado: (estado.usuario is not None and estado.usuario == usuario and (estado.tipo == tipos[0] or estado.tipo == tipos[1])), estados)
    contexto = {'estados': estados_visado}
    return contexto


def ver_documentos_para_visado(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.VISAR)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    if request.method == "POST":
        observacion = request.POST["observaciones"]
        tram = request.POST['tram']
        monto_permiso = tramite.medidas * tramite.tipo_obra.valor_de_superficie
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            if "Aprobar el visado" in request.POST:
                aprobar_visado(request, tram, monto_permiso)
            else:
                no_aprobar_visado(request, tram, observacion)
    else:
        return render(request, 'persona/visador/ver_documentos_tramite.html', {'tramite': tramite,
                                                                               'ctxdoc': documento_set,
                                                                               'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                               "perfil": perfil})
    return redirect('visador')


def ver_documentos_visados(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/visador/ver_documentos_visados.html', contexto)


def aprobar_visado(request, pk_tramite, monto):
    usuario = request.user
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.VISAR, usuario)
    tramite.monto_a_pagar= monto
    tramite.save()
    messages.add_message(request, messages.SUCCESS, 'Tramite visado aprobado')
    return redirect('visador')


def no_aprobar_visado(request, pk_tramite, observacion):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    obs = observacion
    tramite.hacer(tramite.CORREGIR, request.user, obs)
    messages.add_message(request, messages.SUCCESS, 'Tramite con visado no aprobado')
    return redirect('visador')


'''inspector --------------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('inspector')
def mostrar_inspector(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        "ctxtramitesvisadosyconinspeccion": tramites_visados_y_con_inspeccion(),
        "ctxtramitesinspeccionados": tramites_inspeccionados_por_inspector(request),
        "ctxtramitesagendados": tramites_agendados_por_inspector(request)
    }
    for form_name, submit_name in FORMS_INSPECTOR:
        KlassForm = FORMS_INSPECTOR[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/inspector/inspector.html', values)

FORMS_INSPECTOR = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioGrupo,
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def tramites_visados_y_con_inspeccion():
    argumentos = [Visado, Inspeccionado, Aprobado, AprobadoPorPropietario, FinalObraParcialSolicitado,
                  CorreccionesDePrimerInspeccionRealizadas, CorreccionesDeInspeccionRealizadas]
    tramites = Tramite.objects.en_estado(argumentos)
    return tramites


def tramites_inspeccionados_por_inspector(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipos = [9, 12, 19, 22]
    estados_inspeccionados = filter(lambda estado: (estado.usuario is not None and estado.usuario == usuario and
                                                    (estado.tipo == tipos[0] or
                                                     estado.tipo == tipos[1] or
                                                     estado.tipo == tipos[2] or
                                                     estado.tipo == tipos[3])), estados)
    return estados_inspeccionados


def tramites_agendados_por_inspector(request):
    usuario = request.user
    argumentos = [AgendadoPrimerInspeccion, AgendadoInspeccion]
    tramites = Tramite.objects.en_estado(argumentos)
    tramites_del_inspector = filter(lambda t: t.estado().usuario == usuario, tramites)
    tramites_del_inspector_del_dia = filter(lambda t: datetime.datetime.strftime(t.estado().fecha, '%Y-%m-%d') == datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d'), tramites_del_inspector)
    contexto = {'tramites': tramites_del_inspector, 'len_tramites': len(list(tramites_del_inspector_del_dia))}
    return contexto

'''
Metodo que calcula si la hora que quiero agendar un tramite cumple una diferencia de 3 horas con los tramites agendados
es una suposicion sino podria agendar el mismo dia 3 tramites distintos en menos de media hora para un mismo inspector.
'''
def cumple_distancia_en_horas(tramites, fecha):
    horas_de_diferencia = 2
    respuesta = True

    if(len(tramites)>0):
        lista_horas = [t.estado().fecha.hour for t in tramites]
        lista_horas.append(fecha.hour)
    else:
        lista_horas = []

    lista_horas = sorted(lista_horas)

    if (len(lista_horas) > 1):
        for i in range(0,len(lista_horas)-1):
            if lista_horas[i+1] - lista_horas[i] < horas_de_diferencia:
                respuesta = False
                break
    else:
        return True
    return respuesta

def agendar_tramite(request, pk_tramite):
    cant_max_tramites = 3 #Cantidad maxima permitida de tramites a inspeccionar por dia por inspector
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    fecha = parser.parse(request.GET["msg"])
    usuario = request.user
    argumentos = [AgendadoPrimerInspeccion, AgendadoInspeccion]
    tramites = Tramite.objects.en_estado(argumentos)
    tramites_del_inspector = filter(lambda t: t.estado().usuario == usuario, tramites)
    tramites_del_inspector_en_fecha = filter(lambda t: datetime.datetime.strftime(t.estado().fecha, '%Y-%m-%d') == datetime.datetime.strftime(fecha, '%Y-%m-%d'), tramites_del_inspector)
    if(len(list(tramites_del_inspector_en_fecha))<cant_max_tramites and cumple_distancia_en_horas(list(tramites_del_inspector_en_fecha), fecha)):
        tramite.hacer(Tramite.AGENDAR_INSPECCION, request.user, fecha)
        messages.add_message(request, messages.SUCCESS, "La inspeccion ha sido agendada")
    else:
        messages.add_message(request, messages.ERROR, "No es posible asignar mas tramites hasta que no inspeccione al menos uno de sus tramites asignados.")
    return redirect('inspector')


def cargar_inspeccion(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.APROBAR_INSPECCION)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            if "aceptar_tramite" in request.POST:
                aceptar_inspeccion(request, pk_tramite)
            elif "rechazar_tramite" in request.POST:
                rechazar_inspeccion(request, pk_tramite)
    else:
        return render(request, 'persona/inspector/cargar_inspeccion.html', {'tramite': tramite,
                                                                        'ctxdocumentoset': documento_set,
                                                                        'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                        "perfil": perfil})
    return redirect('inspector')


def aceptar_inspeccion(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.APROBAR_INSPECCION, request.user)
    messages.add_message(request, messages.SUCCESS, 'Inspeccion aprobada')
    return redirect('inspector')


def rechazar_inspeccion(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.CORREGIR, request.user, request.POST["observaciones"])
    messages.add_message(request, messages.SUCCESS, 'Inspeccion rechazada')
    return redirect('inspector')


def ver_documentos_tramite_inspector(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/inspector/documentos_tramite_inspector.html', contexto)


#def documentos_inspector_estado(request, pk_estado):
#    usuario = request.user
#    perfil = 'css/' + usuario.persona.perfilCSS
#    estado = get_object_or_404(Estado, pk=pk_estado)
#    fecha = estado.timestamp
#    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
#    documentos = estado.tramite.documentos.all()
#    documentos_fecha = filter(lambda e:(datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
#    contexto= {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
#    return render(request, 'persona/inspector/documentos_del_estado.html', contexto)


'''jefeinspector ----------------------------------------------------------------------------------------'''


@login_required(login_url="login")
@grupo_requerido('jefeinspector')
def mostrar_jefe_inspector(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    values = {
        "perfil": perfil,
        "ctxtramitesconinspeccion": tramite_con_inspecciones_list(),
        "ctxtramitesagendados": tramites_agendados_por_jefeinspector(request),
        "ctxtramitesinspeccionados": tramites_inspeccionados_por_jefeinspector(request),
        "ctxtramitesinspeccionadosporinspectores": tramites_inspeccionados_por_inspectores(),
        "ctxinspectoresconinspeccionesagendadas": inspecciones_agendadas_por_inspectores()
    }
    for form_name, submit_name in FORMS_JEFEINSPECTOR:
        KlassForm = FORMS_JEFEINSPECTOR[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                _form.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/jefe_inspector/jefe_inspector.html', values)

FORMS_JEFEINSPECTOR = {(k.NAME, k.SUBMIT): k for k in [
    FormularioUsuarioGrupo,
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def tramite_con_inspecciones_list():
    argumentos = [FinalObraTotalSolicitado, NoFinalObraTotalSolicitado, CorreccionesDeInspeccionFinalRealizadas]
    tramites = Tramite.objects.en_estado(argumentos)
    contexto = {'tramites': tramites, 'len_tramites': len(tramites)}
    return contexto


def tramites_agendados_por_jefeinspector(request):
    usuario = request.user
    argumentos = [AgendadoInspeccionFinal]
    tramites = Tramite.objects.en_estado(argumentos)
    tramites_del_inspector = filter(lambda t: t.estado().usuario == usuario, tramites)
    dia_hoy = date.today()
    tramites_del_inspector_del_dia = filter(lambda t: datetime.datetime.strftime(t.estado().fecha, '%Y-%m-%d') == datetime.datetime.strftime(dia_hoy, '%Y-%m-%d'), tramites_del_inspector)
    contexto = {'tramites': tramites_del_inspector, 'len_tramites': len(list(tramites_del_inspector_del_dia))}
    return contexto


def tramites_inspeccionados_por_jefeinspector(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipos = [26, 29]
    estados_inspeccionados = filter(lambda estado: (estado.usuario is not None and estado.usuario == usuario and
                                                    (estado.tipo == tipos[0] or estado.tipo == tipos[1])), estados)
    return estados_inspeccionados


def tramites_inspeccionados_por_inspectores():
    estados = Estado.objects.all()
    tipos = [9, 12, 19, 22]
    estados_inspeccionados = filter(lambda estado: (estado.usuario is not None and (estado.tipo == tipos[0] or
                                                                                    estado.tipo == tipos[1] or
                                                                                    estado.tipo == tipos[2] or
                                                                                    estado.tipo == tipos[3])), estados)
    return estados_inspeccionados


def inspectores_sin_inspecciones_agendadas(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    usuarios = Usuario.objects.all()
    inspectores = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(list(lista))):
            if lista[i] == 'inspector':
                if u not in inspectores:
                    inspectores.append(u)
    argumentos = [AgendadoPrimerInspeccion, AgendadoInspeccion]
    tramites = Tramite.objects.en_estado(argumentos)
    estados_ag = []
    for t in tramites:
        estados_ag.append(t.estado())
    estados_agendados = filter(lambda e: e.timestamp.date() == estado.timestamp.date(), estados_ag)
    inspectores_estados_agendados = []
    for i in range(len(list(estados_agendados))):
        inspectores_estados_agendados.append(estados_agendados[i].usuario)
    inspectores_tres_estados_agendados = []
    for i in range(len(list(inspectores_estados_agendados))):
        if inspectores_estados_agendados.count(inspectores_estados_agendados[i]) >= 3:
            inspectores_tres_estados_agendados.append(inspectores_estados_agendados[i])
    inspectores_con_inspeccion_mismo_horario = []
    for i in range(len(list(estados_agendados))):
        if estados_agendados[i].fecha.time() == estado.timestamp.time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
        if estados_agendados[i].fecha.time() < estado.timestamp.time() and estados_agendados[i].fecha.time() >= (estado.timestamp - timedelta(hours=2)).time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
        if estados_agendados[i].fecha.time() > estado.timestamp.time() and estados_agendados[i].fecha.time() <= (estado.timestamp + timedelta(hours=2)).time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
    inspectores_sin_insp_agendadas = []
    for inp in inspectores:
        if inp.is_active and inp not in inspectores_tres_estados_agendados and inp != estado.usuario and inp not in inspectores_con_inspeccion_mismo_horario:
            inspectores_sin_insp_agendadas.append(inp)
    if request.method == "POST" and "cambiar_inspector" in request.POST:
        if request.POST["idusuarioUsuarioS"]:
            pk_inspector = int(request.POST["idusuarioUsuarioS"])
            inspector = Usuario.objects.get(pk=pk_inspector)
            if estado.usuario.id != inspector.id:
                estado.cambiar_usuario(inspector)
                messages.add_message(request, messages.SUCCESS, "El inspector del tramite ha sido cambiado")
            else:
                messages.add_message(request, messages.INFO, "El inspector del tramite no ha sido cambiado. Ha seleccionado el mismo inspector")
        else:
            messages.add_message(request, messages.INFO, "El inspector del tramite no ha sido cambiado. No ha sido seleccionado un inspector")
    else:
        return render(request, 'persona/jefe_inspector/cambiar_inspector_de_inspeccion.html', {'estado': estado, "perfil": perfil, 'inspectores': inspectores_sin_insp_agendadas})
    return redirect('jefeinspector')


def inspecciones_agendadas_por_inspectores():
    argumentos = [AgendadoPrimerInspeccion, AgendadoInspeccion]
    tramites = Tramite.objects.en_estado(argumentos)
    estados_agendados = []
    for t in tramites:
        estados_agendados.append(t.estado())
    return estados_agendados


def agendar_inspeccion_final(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    fecha = parser.parse(request.GET["msg"])
    tramite.hacer(Tramite.AGENDAR_INSPECCION, usuario=request.user, fecha_inspeccion=fecha, inspector=request.user)
    messages.add_message(request, messages.SUCCESS, "La inspeccion final ha sido agendada")
    return redirect('jefeinspector')


def cargar_inspeccion_final(request, pk_tramite):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.APROBAR_INSPECCION)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    id_tramite = int(pk_tramite)
    if request.method == "POST":
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)
        if documento_set.is_valid():
            for docForm in documento_set:
                docForm.save(tramite=tramite)
            if "aceptar_tramite" in request.POST:
                aceptar_inspeccion_final(request, pk_tramite)
            elif "rechazar_tramite" in request.POST:
                rechazar_inspeccion_final(request, pk_tramite)
    else:
        return render(request, 'persona/jefe_inspector/cargar_inspeccion_final.html', {'tramite': tramite,
                                                                            'ctxdocumentoset': documento_set,
                                                                            'documentos_requeridos': tipos_de_documentos_requeridos,
                                                                            "perfil": perfil})
    return redirect('jefeinspector')


def rechazar_inspeccion_final(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.CORREGIR, request.user, request.POST["observaciones"])
    messages.add_message(request, messages.SUCCESS, 'Inspeccion final rechazada')
    return redirect('jefeinspector')


def aceptar_inspeccion_final(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.APROBAR_INSPECCION, request.user)
    messages.add_message(request, messages.SUCCESS, 'Inspeccion final aprobada')
    return redirect('jefeinspector')


def ver_inspecciones(request, pk_tramite):
    pk = int(pk_tramite)
    estados = Estado.objects.all()
    estados_de_tramite = filter(lambda e: (e.tramite.pk == pk), estados)
    estados = filter(lambda e: (e.tipo == 9), estados_de_tramite)
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    contexto = {'estados': estados, "perfil": perfil}
    return render(request, 'persona/jefe_inspector/vista_de_inspecciones.html', contexto)


def ver_documentos_tramite_jefeinspector(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/jefe_inspector/documentos_tramite_jefeinspector.html', contexto)

def ver_documentos_tramite_inspector_por_jefeinspector(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/jefe_inspector/documentos_tramite_inspector_por_jefeinspector.html', contexto)


'''director ---------------------------------------------------------------------------------------------'''

"""
determina que usuarios pueden o no darse de baja logica segun esten relacionados a algun tramite
"""
def usuarios_no_borrables(usuario):

    #Propietario CONSULTAR: nunca se puede dar de baja. se crea a la par de un tramite asiq siempre esta ligado a un tramite en curso.
    #profesional

    #administrativo
    #director: YO lo dejo. poruqe saca todos los directores menos el de la sesion digamos.
    #visador
    #inspector
    #jefe inspector

    estados = Estado.objects.all()
    setattr(usuario, "relacionado", False)
    setattr(usuario, "descripcion", "")

    try:

        if (usuario.persona.profesional):
            tramites = Tramite.objects.filter(profesional=usuario.persona.profesional.id).values_list('id', flat=True)
            if (len(tramites)>0):
                setattr(usuario, "relacionado", True)
                setattr(usuario, "descripcion", "Profesional asignado a tramite: " + ", ".join([str(id_tramite) for id_tramite in tramites]))

        elif(usuario.persona.propietario):
            tramites = Tramite.objects.filter(propietario=usuario.persona.propietario.id).values_list('id', flat=True)
            if (len(tramites)>0):
                setattr(usuario, "relacionado", True)
                setattr(usuario, "descripcion", "Propietario asignado a tramite: " + ", ".join([str(id_tramite) for id_tramite in tramites]))


        elif (usuario.pertenece_a_grupo('visador')):
            tipo = 7
            estados_agendados = list(filter(lambda e: (e.usuario is not None and str(e.tramite.estado()) == "AgendadoParaVisado" and (e.tipo == tipo)), estados))
            if (any(e.usuario.id == usuario.id for e in estados_agendados)):
                setattr(usuario, "relacionado", True)
                setattr(usuario, "descripcion", "Visador asignado a tramite: " +", ".join([str(e.tramite.id) for e in estados_agendados]))

        elif (usuario.pertenece_a_grupo('inspector')):
            tipos = [11, 21, 28]
            estados_agendados= filter(lambda e: (e.usuario is not None and (str(e.tramite.estado()) == 'AgendadoPrimerInspeccion'
                                                or str(e.tramite.estado()) == 'AgendadoInspeccion'
                                                or str(e.tramite.estado()) == 'AgendadoInspeccionFinal')
                                                and (e.tipo == tipos[0] or e.tipo == tipos[1] or e.tipo == tipos[2])), estados)
            if (any(e.usuario.id == usuario.id for e in estados_agendados)):
                setattr(usuario, "relacionado", True)
                setattr(usuario, "descripcion", "Inspector asignado a tramite: " +", ".join([str(e.tramite.id) for e in estados_agendados]))

        elif (usuario.pertenece_a_grupo('jefeinspector')):
            #CHEQUEAR ESTA PARTE VER SI TENGO QUE DARLO DE BAJA O SI TIENE RESTRICCIONES
            setattr(usuario, "relacionado", True)
            setattr(usuario, "descripcion", "Jefe Inspector asignado a un tramite en curso")

    except Exception as e: #Puede originarse por usuarios sin persona
        pass
    return usuario


@login_required(login_url="login")
@grupo_requerido('director')
def mostrar_director(request):
    usuario = request.user
    lista_usuarios = map(usuarios_no_borrables, Usuario.objects.all().exclude(id=request.user.id))
    perfil = 'css/' + usuario.persona.perfilCSS
    tipos_de_documento = TipoDocumento.objects.all()
    values = {
        "lista_usuarios": lista_usuarios,
        "perfil": perfil,
        "datos_usuario": empleados(request.user),
        "tipos_de_documento" : tipos_de_documento,
        "ctxvisadorescontramitesagendados": tramites_con_visado_agendado(),
        "ctxinspectoresconinspeccionesagendadas": inspecciones_agendadas_por_inspectores(),
    }
    for form_name, submit_name in FORMS_DIRECTOR:
        KlassForm = FORMS_DIRECTOR[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST, request.FILES)
            if _form.is_valid():
                try:
                    _form.save()
                    messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                except Exception as e:
                    messages.add_message(request, messages.ERROR, "La accion solicitada no se puede realizar.")
                return redirect(usuario.get_view_name())
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/director/director.html', values)

FORMS_DIRECTOR = {(k.NAME, k.SUBMIT): k for k in [
    FormularioTipoDocumento,
    FormularioUsuarioPersona,  #este formulario no se necesitaria, solo se dan de alta visador, inspector y administrativo
    FormularioTipoObra,
    FormularioTipoDocumento,
    FormularioUsuarioGrupo,
    FormularioUsuarioCambiarDatos,
    FormularioUsuarioContrasenia,
]}


def empleados(director):
    usuarios = Usuario.objects.all().exclude(id=director.id)
    empleados = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(lista)):
            if lista[i] != 'profesional' and lista[i] != 'propietario':
                if u not in empleados:
                    empleados.append(u)
    return empleados


def tramites_con_visado_agendado():
    estados = Estado.objects.all()
    tipos = [7]
    tramites = Tramite.objects.all()
    estados_agendados_para_visado = []
    for tramite in tramites:
        print tramite.estado()
        if tramite.estado().tipo == tipos[0]:
            estados_agendados_para_visado.append(tramite.estado())

    estados_agendados= filter(lambda e: (e.usuario is not None and str(e.tramite.estado()) == 'AgendadoParaVisado' and e.tipo == tipos[0]), estados)
    return estados_agendados_para_visado


def visadores_sin_visado_agendado(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    usuarios = Usuario.objects.all()
    visadores = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(lista)):
            if lista[i] == 'visador':
                if u not in visadores:
                    visadores.append(u)
    #estados = Estado.objects.all()
    #tipo = 7
    #estados_agendados = filter(lambda e: (e.usuario is not None and str(e.tramite.estado()) == 'AgendadoParaVisado' and e.tipo == tipo), estados)
    argumentos = [AgendadoParaVisado]
    tramites = Tramite.objects.en_estado(argumentos)
    estados_agendados = []
    for t in tramites:
        estados_agendados.append(t.estado())
    visadores_estados_agendados = []
    for i in range(len(list(estados_agendados))):
        visadores_estados_agendados.append(estados_agendados[i].usuario)
    print('visadores_estados_agendados:')
    print(visadores_estados_agendados)

    visadores_sin_vis_agendadas = []
    visados_agendados = []
    for vis in visadores:
        visados_agendados = filter(lambda v: vis.pk == v.pk, visadores_estados_agendados)
        print('visadores_agendados:')
        print (visados_agendados)
        if vis != estado.usuario and vis.is_active and len(visados_agendados) < 1 :
            visadores_sin_vis_agendadas.append(vis)
    print('visadores_sin_vis_agendadas:')
    print(visadores_sin_vis_agendadas)
    if request.method == "POST" and "cambiar_visador" in request.POST:
        if request.POST["idusuarioUsuarioS"]:
            pk_visador = int(request.POST["idusuarioUsuarioS"])
            visador = Usuario.objects.get(pk=pk_visador)
            if estado.usuario.id != visador.id:
                estado.cambiar_usuario(visador)
                messages.add_message(request, messages.SUCCESS, "El visador del tramite ha sido cambiado")
            else:
                messages.add_message(request, messages.INFO, "El visador del tramite no ha sido cambiado. Ha seleccionado el mismo visador")
        else:
            messages.add_message(request, messages.INFO, "El visador del tramite no ha sido cambiado. No se ha seleccionado un visador")
    else:
        return render(request, 'persona/director/cambiar_visador_de_tramite.html', {'estado': estado, "perfil": perfil, 'visadores': visadores_sin_vis_agendadas})
    return redirect('director')


def inspectores_sin_inspeccion_agendada(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    usuarios = Usuario.objects.all()
    inspectores = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(list(lista))):
            if lista[i] == 'inspector':
                if u not in inspectores:
                    inspectores.append(u)
    #estados = Estado.objects.all()
    #tipos = [11, 21]
    #estados_agendados = filter(
    #    lambda e: (e.usuario is not None and (str(e.tramite.estado()) == 'AgendadoPrimerInspeccion' or
    #                                          str(e.tramite.estado()) == 'AgendadoInspeccion') and
    #               (e.tipo == tipos[0] or e.tipo == tipos[1])) and
    #              e.timestamp.date() == estado.timestamp.date()
    #    , estados)
    argumentos = [AgendadoPrimerInspeccion, AgendadoInspeccion]
    tramites = Tramite.objects.en_estado(argumentos)
    estados_ag = []
    for t in tramites:
        estados_ag.append(t.estado())
    estados_agendados = filter(lambda e: e.timestamp.date() == estado.timestamp.date(), estados_ag)
    inspectores_estados_agendados = []
    for i in range(len(list(estados_agendados))):
        inspectores_estados_agendados.append(estados_agendados[i].usuario)
    inspectores_tres_estados_agendados = []
    for i in range(len(list(inspectores_estados_agendados))):
        if inspectores_estados_agendados.count(inspectores_estados_agendados[i]) >= 3:
            inspectores_tres_estados_agendados.append(inspectores_estados_agendados[i])
    inspectores_con_inspeccion_mismo_horario = []
    for i in range(len(list(estados_agendados))):
        if estados_agendados[i].fecha.time() == estado.timestamp.time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
        if estados_agendados[i].fecha.time() < estado.timestamp.time() and estados_agendados[i].fecha.time() >= (estado.timestamp - timedelta(hours=2)).time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
        if estados_agendados[i].fecha.time() > estado.timestamp.time() and estados_agendados[i].fecha.time() <= (estado.timestamp + timedelta(hours=2)).time():
            inspectores_con_inspeccion_mismo_horario.append(estados_agendados[i].usuario)
    inspectores_sin_insp_agendadas = []
    for inp in inspectores:
        if inp.is_active and inp not in inspectores_tres_estados_agendados and inp != estado.usuario and inp not in inspectores_con_inspeccion_mismo_horario:
            inspectores_sin_insp_agendadas.append(inp)
    if request.method == "POST" and "cambiar_inspector" in request.POST:
        if request.POST["idusuarioUsuarioS"]:
            print(request.POST["idusuarioUsuarioS"])
            pk_inspector = int(request.POST["idusuarioUsuarioS"])
            print(pk_inspector)
            inspector = Usuario.objects.get(pk=pk_inspector)
            if estado.usuario.id != inspector.id:
                estado.cambiar_usuario(inspector)
                messages.add_message(request, messages.SUCCESS, "El inspector del tramite ha sido cambiado")
            else:
                messages.add_message(request, messages.INFO, "El inspector del tramite no ha sido cambiado. Ha seleccionado el mismo inspector")
        else:
            messages.add_message(request, messages.INFO,
                                 "El inspector del tramite no ha sido cambiado. No se ha seleccionado inspector")
    else:
        return render(request, 'persona/director/cambiar_inspector_d_inspeccion.html', {'estado': estado, "perfil": perfil, 'inspectores': inspectores_sin_insp_agendadas})
    return redirect('director')


def ver_listado_todos_tramites(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS

    argumentos = [Iniciado, ConCorrecciones, ConCorreccionesRealizadas, Aceptado, ConCorreccionesDeVisado,
    CorreccionesDeVisadoRealizadas, AgendadoParaVisado, Visado, ConCorreccionesDePrimerInspeccion,
    CorreccionesDePrimerInspeccionRealizadas, AgendadoPrimerInspeccion, PrimerInspeccion, AprobadoSolicitado,
    Aprobado, NoAprobadoSolicitado, NoAprobado, AprobadoSolicitadoPorPropietario, AprobadoPorPropietario,
    ConCorreccionesDeInspeccion, CorreccionesDeInspeccionRealizadas, AgendadoInspeccion, Inspeccionado,
    FinalObraTotalSolicitado, FinalObraParcialSolicitado, NoFinalObraTotalSolicitado,
    ConCorreccionesDeInspeccionFinal, CorreccionesDeInspeccionFinalRealizadas, AgendadoInspeccionFinal,
    InspeccionFinal, Finalizado, NoFinalizado, FinalObraTotalSolicitadoPorPropietario, Baja]
    lab = ['Iniciado', 'Con Correc.', 'Con Correc. Realizadas', 'Aceptado', 'Con Correc. de Vis.',
    'Correc. Vis. Realizadas', 'Ag.Visado', 'Visado', 'Con Correc. de 1er Insp.',
    'Correc. 1er. Insp. Realizadas', 'Ag. 1er. Inspeccion', '1er Inspeccion', 'Aprob. Sol.',
    'Aprobado', 'No Aprob. Sol.', 'No Aprobado', 'Aprob. Sol. x Prop.', 'Aprob. x Prop.', 'Con Correc. de Insp.',
    'Correc. de Insp. Realizadas', 'Agendado Insp.', 'Inspeccionado', 'F.O.T.S.', 'F.O.P.S.', 'N.F.O.T.S.',
    'Con Correc. de Insp. F.', 'Correc. de Insp. F. Realizadas', 'Ag. Insp. F.',
    'Inspeccion F.', 'Finalizado', 'NoFinalizado', 'F.O.T.S. x Prop.', 'Baja']
    len_argumentos = len(argumentos)
    tramites = Tramite.objects.en_estado(argumentos)
    estados = []
    for t in tramites:
        estados.append(t.estado().tipo)
    estados_cant = dict(collections.Counter(estados))
    for n in range(1, (len_argumentos+1)):
        if (not estados_cant.has_key(n)):
            estados_cant.setdefault(n, 0)
    estados_datos = estados_cant.values()
    cant_est_x_est = dict(zip(lab, estados_datos))
    contexto = {'todos_los_tramites': tramites, "datos_estados": estados_datos, "label_estados": lab, "cant_est_x_est": cant_est_x_est, "perfil": perfil}
    return render(request, 'persona/director/vista_de_tramites.html', contexto)

def reporte_de_tramites_por_tipo(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tipos_obra = TipoObra.objects.all()
    tramites = []
    argumentos_destino = [1 , 2]
    label_destino = ['Vivienda', 'Comercio']
    #destinos = []
    if request.method == "POST":
        if (request.POST.get('id_estado') == '1'):
            estado = Iniciado
        elif (request.POST.get('id_estado') == '2'):
            estado = Finalizado
        elif (request.POST.get('id_estado') == '3'):
            estado = Baja
        # Se filtran tramites por estado
        tramites_estado_requerido = Tramite.objects.en_estado(estado)
        rango_fechas = request.POST.get('daterange')
        fechas = rango_fechas.split(' - ')
        fecha_inicio = datetime.datetime.strptime(fechas[0], "%m/%d/%Y").strftime("%Y-%m-%d")
        fecha_fin = datetime.datetime.strptime(fechas[1], "%m/%d/%Y").strftime("%Y-%m-%d")
        # Se filtra tramites por fecha
        for tramite in tramites_estado_requerido:
            fecha_tramite = tramite.estado().timestamp.date()
            if str(fecha_inicio) <= str(fecha_tramite) <= str(fecha_fin):
                tramites.append(tramite)
        # Se genera rangos de fechas por agrupamiento
        agrupamiento_req = request.POST.get('id_agrupamiento')
        if str(agrupamiento_req) == str(1):
            dias = 1
        if str(agrupamiento_req) == str(2):
            dias = 30
        if str(agrupamiento_req) == str(3):
            dias = 360
        start = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
        end = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')
        step = datetime.timedelta(days=dias)
        lista_dias = []
        while start <= end:
            lista_dias.append(start.date())
            start += step
        if start != end and start < end:
            lista_dias.append(end.date())
        lista_dias.append(start.date())
        rangosLabels = []
        if str(agrupamiento_req) == str(1):
            rangosLabels = lista_dias
        else:
            for i in range(len(lista_dias)):
                if i+1 < len(lista_dias):
                    ini = datetime.datetime.strptime(str(lista_dias[i]), "%Y-%m-%d").strftime("%d-%m-%Y")
                    fin = datetime.datetime.strptime(str(lista_dias[i+1]), "%Y-%m-%d").strftime("%d-%m-%Y")
                    rangosLabels.append(ini + " a " + fin)
        fecha_i = datetime.datetime.strptime(fechas[0], "%m/%d/%Y").strftime("%d-%m-%Y")
        fecha_f = datetime.datetime.strptime(fechas[1], "%m/%d/%Y").strftime("%d-%m-%Y")
        titulosLabels = [estado, fecha_i, fecha_f]
        # Si es destino de obra
        if str(request.POST.get('id_tipo_destino')) == str(1):
            titulosLabels.append('Destino')
            lista_por_fecha_por_destino = {}
            for ld in argumentos_destino:
                listaPorFecha = []
                for i in range(len(lista_dias)):
                    if i + 1 < len(lista_dias):
                        tp = filter(lambda t: t.destino_obra == ld and (str(lista_dias[i]) <= str(t.estado().timestamp.date()) < str(lista_dias[i+1])), tramites)
                        listaPorFecha.append(len(tp))
                lista_por_fecha_por_destino[label_destino[ld-1]] = listaPorFecha
            tram = lista_por_fecha_por_destino
        # Si es tipo de obra
        if str(request.POST.get('id_tipo_destino')) == str(2):
            titulosLabels.append('Tipo')
            lista_por_fecha_por_tipo = {}
            for to in tipos_obra:
                listaPorFecha = []
                for i in range(len(lista_dias)):
                    if i + 1 < len(lista_dias):
                        tp = filter(lambda t: t.tipo_obra == to and (str(lista_dias[i]) <= str(t.estado().timestamp.date()) < str(lista_dias[i + 1])), tramites)
                        listaPorFecha.append(len(tp))
                lista_por_fecha_por_tipo[to] = listaPorFecha
            tram = lista_por_fecha_por_tipo
        contexto = {'todos_los_tramites': tram, 'tramites_tabla': tramites, "perfil": perfil, 'rangosLabels': rangosLabels, 'titulosLabels': titulosLabels}
        return render(request, 'persona/director/reporte_de_tramites_por_tipo.html', contexto)
    else:
        contexto = {"perfil": perfil}
        return render(request, 'persona/director/reporte_de_tramites_por_tipo.html', contexto)


def reporte_de_correciones_profesional(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    tipos_obra = TipoObra.objects.all()
    tramites = []
    tramites_estado_requerido = []
    estados = Estado.objects.all()

    if request.method == "POST" and request.POST.get('buscar'):
        if request.POST.get('id_estado') == '1':
            tipos = 2
            dataset = 'Con correciones'
            tramites_estado_requerido = filter(lambda e: e.tipo == tipos, estados)
        elif request.POST.get('id_estado') == '2':
            tipos = 5
            dataset = 'Con correciones de visado'
            tramites_estado_requerido = filter(lambda e: e.tipo == tipos, estados)
        else:
            tipos = [9, 19, 26]
            dataset = 'Con correciones de inspeccion'
            tramites_estado_requerido = filter(lambda e: (e.tipo == tipos[0] or e.tipo == tipos[1] or e.tipo == tipos[2]), estados)
        #print tramites_estado_requerido
        rango_fechas = request.POST.get('daterange')
        fechas = rango_fechas.split(' - ')
        fecha_inicio = datetime.datetime.strptime(fechas[0], "%m/%d/%Y").strftime("%Y-%m-%d")
        fecha_fin = datetime.datetime.strptime(fechas[1], "%m/%d/%Y").strftime("%Y-%m-%d")
        # Se filtra tramites por fecha
        for tramite in tramites_estado_requerido:
            fecha_tramite = tramite.timestamp.date()
            if str(fecha_inicio) <= str(fecha_tramite) <= str(fecha_fin):
                tramites.append(tramite)
        # Se genera rangos de fechas por agrupamiento
        agrupamiento_req = request.POST.get('id_agrupamiento')
        if str(agrupamiento_req) == str(1):
            dias = 1
        if str(agrupamiento_req) == str(2):
            dias = 30
        if str(agrupamiento_req) == str(3):
            dias = 360
        start = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
        end = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')
        step = datetime.timedelta(days=dias)
        lista_dias = []
        while start <= end:
            lista_dias.append(start.date())
            start += step
        if start != end and start < end:
            lista_dias.append(end.date())
        lista_dias.append(start.date())
        rangosLabels = []
        for i in range(len(lista_dias)):
            rangosLabels.append(i)
        listaPorFecha = []
        for i in range(len(lista_dias)):
            if i + 1 < len(lista_dias):
                tp = filter(lambda t: str(lista_dias[i]) <= str(t.timestamp.date()) < str(lista_dias[i + 1]), tramites)
                listaPorFecha.append(len(tp))
        tram = listaPorFecha
        fecha_i = datetime.datetime.strptime(fechas[0], "%m/%d/%Y").strftime("%d-%m-%Y")
        fecha_f = datetime.datetime.strptime(fechas[1], "%m/%d/%Y").strftime("%d-%m-%Y")
        datos_titulo = [dataset, fecha_i, fecha_f]

        #linea de tendencia
        linea_tendencia = []
        if fecha_inicio < fecha_fin:
            t1 = 0.0
            t2 = 0.0
            y1 = 0.0
            y2 = 0.0
            for i in range(len(tram)):
                if i <= (len(tram)/2):
                    y1 =y1 + tram[i]
                else:
                    y2 = y2 + tram[i]
            y1 = y1/ (len(tram)/2)
            y2 = y2 / (len(tram) / 2)
            for i in range(len(rangosLabels)):
                if i <= (len(rangosLabels)/2):
                    t1 =t1 + rangosLabels[i]
                else:
                    t2 = t2 + rangosLabels[i]
            t1 = t1 / (len(rangosLabels) / 2)
            t2 = t2 / (len(rangosLabels) / 2)
            x = 0
            x1 = len(tram)
            y = 0
            y1 = 0
            y = ((y2 - y1)/(t2 - t1))*( x * t1) + y1
            y1 = ((y2 - y1) / (t2 - t1)) * (x1 * t1) + y1

            linea_tendencia.append(int(x))
            linea_tendencia.append(int(y))
            linea_tendencia.append(int(x1))
            linea_tendencia.append(int(y1))

            print ("-------------------------")
            print linea_tendencia

        contexto = {'todos_los_tramites': tram, 'tramites_tabla': tramites, "perfil": perfil, 'rangosLabels': rangosLabels, 'dataset': dataset, 'datos_titulo': datos_titulo, 'linea_tendencia': linea_tendencia}
        return render(request, 'persona/director/reporte_de_correcciones.html', contexto)
    else:
        contexto = {"perfil": perfil}
        return render(request, 'persona/director/reporte_de_correcciones.html', contexto)


def empleados_con_director():
    usuarios = Usuario.objects.all()
    empleados = []
    for u in usuarios:
        lista = list(u.groups.values_list('name', flat=True))
        for i in range(len(lista)):
            if lista[i] != 'profesional' and lista[i] != 'propietario':
                if u not in empleados:
                    empleados.append(u)
    return empleados


def ver_listado_todos_usuarios(request):
    grupos = Group.objects.all()
    label_grupos = []
    for g in grupos:
        if g.name != 'profesional' and g.name != 'propietario':
            label_grupos.append(g.name)
    usuarios = Usuario.objects.all()
    cant_usuarios_grupos = []
    for u in usuarios:
        for gu in u.get_view_groups():
            if str(gu) != 'profesional' and str(gu) != 'propietario':
                cant_usuarios_grupos.append(gu)
    total_usuarios_grupos = dict(collections.Counter(cant_usuarios_grupos))
    for lg in label_grupos:
        if not total_usuarios_grupos.has_key(lg):
            total_usuarios_grupos.setdefault(lg, 0)
    datos_grupos = total_usuarios_grupos.values()
    usuarios = empleados_con_director()
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    return render(request,
                  'persona/director/vista_de_usuarios.html',
                  {'todos_los_usuarios': usuarios, "label_grupos": label_grupos,
                   "datos_grupos": datos_grupos, "perfil": perfil})


def ver_actividad_usuario(request, usuario):
    usuarios = Usuario.objects.all()
    usuario_req = filter(lambda u: (str(u) == usuario), usuarios)
    estados = Estado.objects.all()
    estados_usuario_req = filter(lambda estado: (str(estado.usuario) == str(usuario_req[0])), estados)
    fechas_del_estado = []
    estados_por_fecha = []
    for e in estados_usuario_req:
        estados_por_fecha.append(e.timestamp.strftime("%d/%m/%Y"))
        if (e.timestamp.strftime("%d/%m/%Y")) not in fechas_del_estado:
            fechas_del_estado.append(e.timestamp.strftime("%d/%m/%Y"))
    cant_estados_por_fecha = [estados_por_fecha.count(f) for f in fechas_del_estado]
    user = request.user
    perfil = 'css/' + user.persona.perfilCSS
    return render(request, 'persona/director/vista_de_actividad_usuario.html', {"perfil": perfil,
                                                                                "usuario": usuario_req[0],
                                                                                "estados": estados_usuario_req,
                                                                                "fechas_estados": fechas_del_estado,
                                                                                "cant_estados": cant_estados_por_fecha})


def detalle_de_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    pk = int(pk_tramite)
    estados = Estado.objects.all()
    estados_de_tramite = filter(lambda e: (e.tramite.pk == pk), estados)
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = []
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"))
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    return render(request, 'persona/director/detalle_de_tramite.html', {"tramite": contexto0, "estados": contexto1, "fecha": fechas_del_estado, "perfil": perfil})


def ver_documentos_del_estado(request, pk_estado):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = datetime.datetime.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e:(datetime.datetime.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto= {'documentos_de_fecha': documentos_fecha, "perfil": perfil}
    return render(request, 'persona/director/ver_documentos_del_estado.html', contexto)


"""
Metodo que se encarga de devolver todos los profesionales con usuario
Se utiliza en la vista de director
"""
@login_required(login_url="login")
@grupo_requerido('director')
def listado_profesionales_director(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    personas = Persona.objects.all()
    profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
    contexto = {'profesionales': profesionales_con_usuario, "perfil": perfil}
    return render(request, 'persona/director/profesional_list_con_usuario_director.html', contexto)


"""
Metodo que se encarga de devolver todos los propietarios con usuario
Se utiliza en la vista de director
"""
@login_required(login_url="login")
@grupo_requerido('director')
def listado_propietarios_director(request):
    usuario = request.user
    perfil = 'css/' + usuario.persona.perfilCSS
    propietarios = Propietario.objects.all()
    propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None ), propietarios)
    contexto = {'propietarios': propietarios_con_usuario, "perfil": perfil}
    return render(request, 'persona/director/propietario_list_con_usuario_director.html', contexto)


"""
Se encarga de devolver todos los profesionales con usuario en una archivo de excel
Se utiliza en la vista de director
"""
class ReporteProfesionalesDirectorExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        cont = 0
        personas = Persona.objects.all()
        profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'lISTADO DE PROFESIONALES'
        ws.merge_cells('B1:K1')
        ws['B2'] = 'NOMBRE'
        ws['C2'] = 'APELLIDO'
        ws['D2'] = 'MAIL'
        ws['E2'] = 'DIRECCION'
        ws['F2'] = 'CUIL'
        ws['G2'] = 'TELEFONO'
        ws['H2'] = 'PROFESION'
        ws['I2'] = 'CATEGORIA'
        ws['J2'] = 'MATRICULA'
        cont = 3
        for p in profesionales_con_usuario:
            ws.cell(row=cont, column=2).value = str(p.nombre)
            ws.cell(row=cont, column=3).value = str(p.apellido)
            ws.cell(row=cont, column=4).value = str(p.mail)
            ws.cell(row=cont, column=5).value = str(p.domicilio_persona)
            ws.cell(row=cont, column=6).value = str(p.cuil)
            ws.cell(row=cont, column=7).value = str(p.telefono)
            ws.cell(row=cont, column=8).value = str(p.profesional.profesion)
            ws.cell(row=cont, column=9).value = str(p.profesional.categoria)
            ws.cell(row=cont, column=10).value = str(p.profesional.matricula)
            cont = cont + 1
        nombre_archivo = "ListadoProfesionalesDirectorExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


"""
Se encarga de devolver todos los propietarios con usuario en una archivo de excel
Se utiliza en la vista de director
"""
class ReportePropietariosDirectorExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        cont = 0
        propietarios = Propietario.objects.all()
        propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None), propietarios)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LISTADO DE PROPIETARIOS'
        ws.merge_cells('B1:G1')
        ws['B2'] = 'NOMBRE'
        ws['C2'] = 'APELLIDO'
        ws['D2'] = 'CUIL'
        ws['E2'] = 'TELEFONO'
        ws['F2'] = 'EMAIL'
        cont = 3
        for p in propietarios_con_usuario:
            ws.cell(row=cont, column=2).value = str(p.persona.nombre)
            ws.cell(row=cont, column=3).value = str(p.persona.apellido)
            ws.cell(row=cont, column=4).value = str(p.persona.cuil)
            ws.cell(row=cont, column=5).value = str(p.persona.telefono)
            ws.cell(row=cont, column=6).value = str(p.persona.mail)
            cont = cont + 1
        nombre_archivo = "ListadoPropietariosDirectorExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


"""
Se encarga de devolver todos los profesionales con usuario en un archivo de pdf
Se utiliza en la vista de director
"""
class ReporteProfesionalesDirectorPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Listado de profesionales Director.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )
        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))
        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)
        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Listado de profesionales'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))
        encabezados = ('NOMBRE', 'APELLIDO', 'MAIL', 'CUIL', 'TELEFONO', 'PROFESION', 'CAT.', 'MAT.')
        personas = Persona.objects.all()
        profesionales_con_usuario = filter(lambda persona: (persona.usuario is not None and persona.profesional is not None), personas)
        detalles = [
            (p.nombre, p.apellido, p.mail, p.cuil, p.telefono, p.profesional.profesion, p.profesional.categoria, p.profesional.matricula)
            for p in profesionales_con_usuario]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 4 * cm, 3 * cm, 2 * cm, 3.5 * cm, 1 * cm, 2.5 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response

"""
Se encarga de devolver todos los propietarios con usuario en un archivo de pdf
Se utiliza en la vista de director
"""
class ReportePropietariosDirectorPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Listado de propietarios Director.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )
        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))
        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)
        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Listado de propietarios'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))
        encabezados = ('NOMBRE', 'APELLIDO', 'CUIL', 'TELEFONO', 'EMAIL')
        propietarios = Propietario.objects.all()
        propietarios_con_usuario = filter(lambda propietario: (propietario.persona.usuario is not None and propietario.persona is not None),propietarios)
        detalles = [(p.persona.nombre, p.persona.apellido, p.persona.cuil, p.persona.telefono, p.persona.mail)
            for p in propietarios_con_usuario]
        detalle_orden = Table([encabezados] + detalles, colWidths=[3 * cm, 3 * cm, 3 * cm, 3 * cm, 5 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response


class ReporteTramitesDirectorExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LISTADO DE TRAMITES'
        ws.merge_cells('B1:G1')
        ws['B2'] = 'NRO'
        ws['C2'] = 'TIPO_DE_OBRA'
        ws['D2'] = 'PROFESIONAL'
        ws['E2'] = 'PROPIETARIO'
        ws['F2'] = 'MEDIDAS'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = tramite.id
            ws.cell(row=cont, column=3).value = str(tramite.tipo_obra)
            ws.cell(row=cont, column=4).value = str(tramite.profesional)
            ws.cell(row=cont, column=5).value = str(tramite.propietario)
            ws.cell(row=cont, column=6).value = tramite.medidas
            cont = cont + 1
        nombre_archivo = "ListadoTramitesExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteTramitesDirectorPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Listado de tramites.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )
        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))
        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)
        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Listado de tramites'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))

        encabezados = ('NRO', 'TIPO', 'PROFESIONAL', 'PROPIETARIO', 'M2', 'ESTADO')
        detalles = [
            (tramite.id, tramite.tipo_obra, tramite.profesional, tramite.propietario, tramite.medidas, tramite.estado())
            for tramite in
            Tramite.objects.all()]
        detalle_orden = Table([encabezados] + detalles, colWidths=[1 * cm, 5 * cm, 4 * cm, 4 * cm, 1 * cm, 5.5 * cm])

        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response


class ReporteEmpleadosDirectorExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        usuarios = empleados_con_director()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LISTADO DE EMPLEADOS'
        ws.merge_cells('B1:H1')
        ws['B2'] = 'USUARIO'
        ws['C2'] = 'GRUPO'
        ws['D2'] = 'NOMBRE'
        ws['E2'] = 'DOCUMENTO'
        ws['F2'] = 'TELEFONO'
        ws['G2'] = 'MAIL'
        cont = 3
        for u in usuarios:
            ws.cell(row=cont, column=2).value = str(u)
            ws.cell(row=cont, column=3).value = str(u.get_view_groups()[0])
            ws.cell(row=cont, column=4).value = str(u.persona)
            ws.cell(row=cont, column=5).value = str(u.persona.dni)
            ws.cell(row=cont, column=6).value = str(u.persona.telefono)
            ws.cell(row=cont, column=7).value = str(u.persona.mail)
            cont = cont + 1

        nombre_archivo = "ListadoEmpleadosExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class RotarImagen(Image):
    def wrap(self, availWidth, availHeight):
        self.w, self.h = Image.wrap(self, availWidth, availHeight)
        return self.w, self.h

    def draw(self):
        self.canv.translate(0,self.h)
        self.canv.rotate(-90)
        Image.draw(self)


class ReporteEmpleadosDirectorPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Informe de empleados.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )

        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))

        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))

        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)

        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Informe de empleados'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))

        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))

        encabezados = ('USUARIO', 'GRUPO', 'NOMBRE', 'DOCUMENTO ', 'TELEFONO', 'MAIL')
        detalles = []
        detalles = [
            (u, u.get_view_groups()[0], u.persona,u.persona.dni, u.persona.telefono, u.persona.mail)
            for u in empleados_con_director()]

        detalle_orden = Table([encabezados] + detalles, colWidths=[4 * cm, 3 * cm, 4 * cm, 2 * cm, 2 * cm, 4 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response


'''general ----------------------------------------------------------------------------------------------'''

def cambiar_perfil(request):
    usuario = request.user
    if request.method == "POST":
        estilo = request.POST["estiloCSS"]
        usuario.persona.modificarPerfilCSS(estilo)
        return redirect(usuario.get_view_name())


'''Funcionalidades 2018 pre-final -----------------------------------------------------------------------'''

def alta_persona(request):
    if request.method == "POST":
        form = FormularioPersona(request.POST)
        if form.is_valid():
            persona = form.save()
            persona.save()
    else:
        form = FormularioPersona()
    return render(request, 'persona/alta/alta_persona.html', {'form': form})


"""
Metodo que se encarga de realizar el alta baja de los usuarios
Se utiliza en la vista de director
"""
@login_required(login_url="login")
@grupo_requerido('director')
def alta_baja_usuarios(request):
    if request.method == 'POST':
        id_usuario = request.POST.get('id_form_usuario')
        opcion = request.POST.get('id_form_opcion')
        usuario = Usuario.objects.get(id=int(id_usuario))
        if (opcion == 'Baja'):
            usuario.is_active = False
        else:
            usuario.is_active = True
        usuario.save()
    return redirect('director')


"""
Metodo que se encarga de devolver los grupos a los que pertenece un usuario
Se utiliza en la vista de director
"""
@login_required(login_url="login")
@grupo_requerido('director')
def get_grupos_usuario(request):
    if request.is_ajax():
        id = request.GET.get('usuario_id')
        lista = Usuario.objects.filter(id=int(id)).values_list('groups__name',flat=True)
        return JsonResponse(list(lista), safe=False)


def cambiar_descrip_filas(columna):
    col = str(columna)
    if ('7' in col):
        return 'Agendado para visado'
    elif ('8' in col):
        return 'Visado'
    elif ('5' in col):
        return "Con correcciones de visado"
    return columna

def boxplot(request):
    contexto={}
    lista=[]

    lista_visadores = Usuario.objects.filter(groups__name='visador')
    lista_inspectores = Usuario.objects.filter(groups__name='inspector')
    contexto['lista_inspectores']=lista_inspectores
    contexto['lista_visadores']=lista_visadores

    parametros = [7,8,5]
    usuarios = [lista_visadores.first().id]
    contexto['boton_presionado'] = usuarios[0]

    if 'boton_plotbox' in request.POST:
        if 'opciones' in request.POST:
            if 'todos_visadores' in request.POST.get('opciones'):
                parametros = [7,8,5]          #VISADOR AgendadoParaVisado, Visado, ConCorreccionesDeVisado
                usuarios = Usuario.objects.filter(groups__name='visador').values_list('id',flat=True)

            elif 'todos_inspectores' in request.POST.get('opciones'):
                parametros= [11,12,9]         #INSPECTOR AgendadoPrimerInspeccion, PrimerInspeccion, ConCorreccionesDePrimerInspeccion
                usuarios = Usuario.objects.filter(groups__name='inspector').values_list('id',flat=True)

            else:
                id = request.POST.get('opciones')
                u = Usuario.objects.get(id=int(id))
                usuarios = [u.id]
                if u.get_view_name() == 'inspector':
                    parametros = [11,12,9]
                else:
                    parametros = parametros = [7,8,5]
            contexto['boton_presionado'] = request.POST.get('opciones')

    tramites_agendados = Estado.objects.filter(tipo__in=parametros, usuario__in=usuarios).values('usuario__username','tramite_id','timestamp','tipo').order_by('tramite_id','timestamp')
    df_tramites = pd.DataFrame.from_records(tramites_agendados)
    if not df_tramites.empty:
        tramite_id = df_tramites.groupby(['tramite_id']).timestamp.count().items()
        tramites_a_borrar = [x[0] for x in tramite_id if x[1]%2!=0]

        if tramites_a_borrar:
            for t in reversed(tramites_a_borrar): #la tengo que dar vuelta sino los indices no coinciden
                indice = df_tramites[(df_tramites.tramite_id == t)].index.max()
                df_tramites = df_tramites.drop(df_tramites.index[indice])

        df=pd.pivot_table(df_tramites,index=['tramite_id','timestamp','usuario__username'], values='tipo', aggfunc='first', fill_value=0).reset_index()
        df['tipo'] = df['tipo'].apply(cambiar_descrip_filas)
        df['timestamp'] = df['timestamp'].apply(lambda row: row.strftime('%d/%m/%Y'))
        for nombre in df.usuario__username.unique():
            columna_temporal =  df[(df.usuario__username == nombre)].timestamp
            pares = [datetime.datetime.strptime(x,'%d/%m/%Y').date() for x in columna_temporal[::2]]
            impares = [datetime.datetime.strptime(y,'%d/%m/%Y').date() for y in columna_temporal[1::2]]
            resta = [t.days for t in list(map(operator.sub, impares,pares))]
            lista.append({nombre:resta})

        df=df.to_html(index=False, classes=["table table-condensed", "table-bordered", "table-striped", "table-hover"])
        contexto['lista']=lista
        contexto['df']=df

    else:
        messages.add_message(request, messages.WARNING, "No existen datos disponibles para la consulta")

    return render(request, 'persona/director/reporte_boxplot.html', contexto)


def generar_boxplot(request):
    import plotly.offline as offline
    from selenium import webdriver
    import plotly.plotly as py
    import plotly.graph_objs as go
    import numpy as np

    contexto={}
    lista=[]
    data=[]

    lista_visadores = Usuario.objects.filter(groups__name='visador')
    lista_inspectores = Usuario.objects.filter(groups__name='inspector')
    contexto['lista_inspectores']=lista_inspectores
    contexto['lista_visadores']=lista_visadores

    parametros = [7,8,5]
    usuarios = [lista_visadores.first().id]

    opcion = request.get('opcion')

    if opcion:
        if 'todos_visadores' == opcion:
            parametros = [7,8,5]          #VISADOR AgendadoParaVisado, Visado, ConCorreccionesDeVisado
            usuarios = Usuario.objects.filter(groups__name='visador').values_list('id',flat=True)

        elif 'todos_inspectores' == opcion:
            parametros= [11,12,9]         #INSPECTOR AgendadoPrimerInspeccion, PrimerInspeccion, ConCorreccionesDePrimerInspeccion
            usuarios = Usuario.objects.filter(groups__name='inspector').values_list('id',flat=True)

        else:
            id = opcion
            u = Usuario.objects.get(id=int(id))
            usuarios = [u.id]
            if u.get_view_name() == 'inspector':
                parametros = [11,12,9]
            else:
                parametros = parametros = [7,8,5]

    #Genero el DataFrame con los datos iniciales
    tramites_agendados = Estado.objects.filter(tipo__in=parametros, usuario__in=usuarios).values('usuario__username','tramite_id','timestamp','tipo').order_by('tramite_id','timestamp')
    df_tramites = pd.DataFrame.from_records(tramites_agendados)
    if not df_tramites.empty:
        tramite_id = df_tramites.groupby(['tramite_id']).timestamp.count().items()
        tramites_a_borrar = [x[0] for x in tramite_id if x[1]%2!=0]

        if tramites_a_borrar:
            for t in reversed(tramites_a_borrar): #la tengo que dar vuelta sino los indices no coinciden
                indice = df_tramites[(df_tramites.tramite_id == t)].index.max()
                df_tramites = df_tramites.drop(df_tramites.index[indice])

        df=pd.pivot_table(df_tramites,index=['tramite_id','timestamp','usuario__username'], values='tipo', aggfunc='first', fill_value=0).reset_index()
        df['tipo'] = df['tipo'].apply(cambiar_descrip_filas)
        df['timestamp'] = df['timestamp'].apply(lambda row: row.strftime('%d/%m/%Y'))
        for nombre in df.usuario__username.unique():
            columna_temporal =  df[(df.usuario__username == nombre)].timestamp
            pares = [datetime.datetime.strptime(x,'%d/%m/%Y').date() for x in columna_temporal[::2]]
            impares = [datetime.datetime.strptime(y,'%d/%m/%Y').date() for y in columna_temporal[1::2]]
            resta = [t.days for t in list(map(operator.sub, impares,pares))]
            lista.append({nombre:resta})


    #Genero el boxplot
    for dic in lista:
        for k,v in dic.items():
            data.append(go.Box(name=str(k), y=v))
    fig = go.Figure(data=data)

    #Genero el html que me da la imagen
    offline.plot(fig, image='svg', auto_open=False, image_width=1000, image_height=500)

    #Selenium y PhantomJS para guardar la imagen
    driver = webdriver.PhantomJS()
    driver.set_window_size(1000, 500)
    driver.get('temp-plot.html')
    driver.save_screenshot(settings.MEDIA_ROOT + '/boxplot.png')

    return df



from reportlab.platypus import SimpleDocTemplate, Image
class boxplot_to_pdf(View):

    def get(self, request, *args, **kwargs):
        df = generar_boxplot(kwargs)
        filename = "Reporte_productividad_empleados.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0,
            leftMargin=0,
            topMargin=0,
            bottomMargin=0,
        )

        story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=10))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))

        usuario = 'Usuario: ' + str(request.user.persona) + ' -  Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        story.append(Paragraph(usuario, styles["Usuario"]))
        story.append(Spacer(0, cm * 0.15))

        im1 = Image(settings.MEDIA_ROOT + '/imagenes/banner_municipio_3.png', width=630, height=50)
        im1.hAlign = 'CENTER'
        story.append(im1)

        story.append(Spacer(0, cm * 0.05))
        subtitulo = 'Reporte de empleados'
        story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        story.append(Spacer(0, cm * 0.15))

        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=640, height=3)
        story.append(im0)
        story.append(Spacer(0, cm * 0.5))

        boxplot = Image(settings.MEDIA_ROOT + '/boxplot.png', width=400, height=250)
        story.append(boxplot)

        story.append(Spacer(0, cm * 0.5))

        encabezados = ('TRAMITE', 'TIMESTAMP', 'NOMBRE DE USUARIO', 'TIPO ')
        detalles = df.values.tolist()

        detalle_orden = Table([encabezados] + detalles, colWidths=[1 * cm, 3 * cm, 4 * cm, 4 * cm, 2 * cm, 3 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        story.append(detalle_orden)
        doc.build(story)
        return response

class boxplot_to_excel(View):
    def get(self, request, *args, **kwargs):
        try:
            from io import BytesIO as IO # for modern python
        except ImportError:
            from StringIO import StringIO as IO # for legacy python

        lista_visadores = Usuario.objects.filter(groups__name='visador')
        lista_inspectores = Usuario.objects.filter(groups__name='inspector')
        parametros = [7,8,5]
        usuarios = [lista_visadores.first().id]

        opcion = kwargs.get('opcion')
        if opcion:
            if 'todos_visadores' == opcion:
                parametros = [7,8,5]          #VISADOR AgendadoParaVisado, Visado, ConCorreccionesDeVisado
                usuarios = Usuario.objects.filter(groups__name='visador').values_list('id',flat=True)

            elif 'todos_inspectores' == opcion:
                parametros= [11,12,9]         #INSPECTOR AgendadoPrimerInspeccion, PrimerInspeccion, ConCorreccionesDePrimerInspeccion
                usuarios = Usuario.objects.filter(groups__name='inspector').values_list('id',flat=True)

            else:
                id = opcion
                u = Usuario.objects.get(id=int(id))
                usuarios = [u.id]
                if u.get_view_name() == 'inspector':
                    parametros = [11,12,9]
                else:
                    parametros = parametros = [7,8,5]

        #Genero el DataFrame con los datos iniciales
        tramites_agendados = Estado.objects.filter(tipo__in=parametros, usuario__in=usuarios).values('usuario__username','tramite_id','timestamp','tipo').order_by('tramite_id','timestamp')
        df_tramites = pd.DataFrame.from_records(tramites_agendados)
        if not df_tramites.empty:
            tramite_id = df_tramites.groupby(['tramite_id']).timestamp.count().items()
            tramites_a_borrar = [x[0] for x in tramite_id if x[1]%2!=0]

            if tramites_a_borrar:
                for t in reversed(tramites_a_borrar): #la tengo que dar vuelta sino los indices no coinciden
                    indice = df_tramites[(df_tramites.tramite_id == t)].index.max()
                    df_tramites = df_tramites.drop(df_tramites.index[indice])

            df=pd.pivot_table(df_tramites,index=['tramite_id','timestamp','usuario__username'], values='tipo', aggfunc='first', fill_value=0).reset_index()
            df['tipo'] = df['tipo'].apply(cambiar_descrip_filas)
            df['timestamp'] = df['timestamp'].apply(lambda row: row.strftime('%d/%m/%Y'))


        excel_file = IO()
        xlwriter = pd.ExcelWriter(excel_file, engine='xlsxwriter')
        df.set_index(df.columns[0], inplace=True)
        df.to_excel(xlwriter, 'sheetname')
        xlwriter.save()
        xlwriter.close()

        # important step, rewind the buffer or when it is read() you'll get nothing
        # but an error message when you try to open your zero length file in Excel
        excel_file.seek(0)

        # set the mime type so that the browser knows what to do with the file
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # set the file name in the Content-Disposition header
        response['Content-Disposition'] = 'attachment; filename=Reporte_productividad_empleados.xlsx'

        return response
